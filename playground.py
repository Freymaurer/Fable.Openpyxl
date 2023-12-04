from __future__ import annotations
from abc import abstractmethod
import openpyxl as openpyxl_1
from typing import (Any, Protocol, Callable)
from fable_modules.fable_library.reflection import (TypeInfo, union_type)
from fable_modules.fable_library.string_ import (to_console, printf)
from fable_modules.fable_library.types import (Array, Union)
from fable_modules.fable_library.util import (IEnumerable_1, IEnumerable, equals)

def _expr0() -> TypeInfo:
    return union_type("Playground.CellType", [], CellType, lambda: [[], [], [], [], [], []])


class CellType(Union):
    def __init__(self, tag: int, *fields: Any) -> None:
        super().__init__()
        self.tag: int = tag or 0
        self.fields: Array[Any] = list(fields)

    @staticmethod
    def cases() -> list[str]:
        return ["Float", "Integer", "String", "Boolean", "DateTime", "Empty"]


CellType_reflection = _expr0

def CellType_fromCellType_Z721C83C5(cell_type: str) -> CellType:
    if cell_type == "int":
        return CellType(1)

    elif cell_type == "float":
        return CellType(0)

    elif cell_type == "str":
        return CellType(2)

    elif cell_type == "bool":
        return CellType(3)

    elif cell_type == "datetime":
        return CellType(4)

    elif cell_type == "NoneType":
        return CellType(5)

    else: 
        raise Exception(("Unknown cell type of type: \'" + cell_type) + "\'")



class Cell(Protocol):
    @property
    @abstractmethod
    def cell_type(self) -> str:
        ...

    @property
    @abstractmethod
    def value(self) -> Any:
        ...

    @value.setter
    @abstractmethod
    def value(self, __arg0: Any) -> None:
        ...


class Table(Protocol):
    @property
    @abstractmethod
    def display_name(self) -> str:
        ...

    @display_name.setter
    @abstractmethod
    def display_name(self, __arg0: str) -> None:
        ...

    @property
    @abstractmethod
    def header_row_count(self) -> bool:
        ...

    @header_row_count.setter
    @abstractmethod
    def header_row_count(self, __arg0: bool) -> None:
        ...

    @property
    @abstractmethod
    def id(self) -> int:
        ...

    @id.setter
    @abstractmethod
    def id(self, __arg0: int) -> None:
        ...

    @property
    @abstractmethod
    def name(self) -> str:
        ...

    @name.setter
    @abstractmethod
    def name(self, __arg0: str) -> None:
        ...


class TableMap(Protocol):
    @abstractmethod
    def Item(self, __arg0: str) -> Table:
        ...

    @abstractmethod
    def delete(self, displayName: str) -> None:
        ...

    @abstractmethod
    def items(self) -> Array[tuple[str, str]]:
        ...

    @abstractmethod
    def values(self) -> Array[Table]:
        ...


class Worksheet(Protocol):
    @abstractmethod
    def add_table(self, __arg0: Table) -> None:
        ...

    @abstractmethod
    def append(self, __arg0: Array[Any]) -> None:
        ...

    @abstractmethod
    def delete_cols(self, start_index: int, count: int) -> None:
        ...

    @abstractmethod
    def delete_rows(self, start_index: int, count: int) -> None:
        ...

    @abstractmethod
    def delete_table(self, displayName: str) -> None:
        ...

    @property
    @abstractmethod
    def columns(self) -> Array[Array[Cell]]:
        ...

    @property
    @abstractmethod
    def rows(self) -> Array[Array[Cell]]:
        ...

    @property
    @abstractmethod
    def table_count(self) -> int:
        ...

    @property
    @abstractmethod
    def tables(self) -> TableMap:
        ...

    @property
    @abstractmethod
    def title(self) -> str:
        ...

    @title.setter
    @abstractmethod
    def title(self, __arg0: str) -> None:
        ...

    @property
    @abstractmethod
    def values(self) -> Array[Array[Any]]:
        ...

    @abstractmethod
    def insert_cols(self, __arg0: int) -> None:
        ...

    @abstractmethod
    def insert_rows(self, __arg0: int) -> None:
        ...

    @abstractmethod
    def iter_cols(self, min_row: int, max_col: int, max_row: int, action: Callable[[Array[Cell]], None]) -> None:
        ...

    @abstractmethod
    def iter_rows(self, min_row: int, max_col: int, max_row: int, action: Callable[[Array[Cell]], None]) -> None:
        ...


class Workbook(IEnumerable_1, IEnumerable[Any]):
    @abstractmethod
    def Item(self, __arg0: str) -> Worksheet:
        ...

    @abstractmethod
    def copy_worksheet(self, __arg0: Worksheet) -> Worksheet:
        ...

    @abstractmethod
    def create_sheet(self, __arg0: str, position: int | None) -> Worksheet:
        ...

    @property
    @abstractmethod
    def active(self) -> Worksheet:
        ...

    @property
    @abstractmethod
    def iso_dates(self) -> bool:
        ...

    @iso_dates.setter
    @abstractmethod
    def iso_dates(self, __arg0: bool) -> None:
        ...

    @property
    @abstractmethod
    def sheetnames(self) -> Array[str]:
        ...

    @property
    @abstractmethod
    def template(self) -> bool:
        ...

    @template.setter
    @abstractmethod
    def template(self, __arg0: bool) -> None:
        ...


class BytesIO(Protocol):
    @abstractmethod
    def ToFile(self, path: str) -> None:
        ...

    @property
    @abstractmethod
    def x(self) -> str:
        ...

    @abstractmethod
    def getbuffer(self) -> Any:
        ...

    @abstractmethod
    def getvalue(self) -> bytearray:
        ...


class WorkbookStatic(Protocol):
    @abstractmethod
    def create(self) -> Workbook:
        ...


class TableStatic(Protocol):
    @abstractmethod
    def create(self, displayName: str, ref: str) -> Table:
        ...


class BytesIOStatic(Protocol):
    pass

class OpenPyXL(Protocol):
    @abstractmethod
    def Workbook(self) -> Workbook:
        ...

    @abstractmethod
    def load_workbook(self, __arg0: str) -> Workbook:
        ...


def write_bytes(bytes: bytearray, path: str) -> None:
    
  # Write the bytes data to the output file path using shutil
  with open(path, 'wb') as output_file:
      output_file.write(bytes)



openpyxl: OpenPyXL = openpyxl_1

path_to_file: str = "tests\\TestFiles\\MinimalTest.xlsx"

wb: Workbook = openpyxl.load_workbook(path_to_file)

ws: Worksheet = wb.active

cell: Cell = ws.cell(1, 1)

arg: bool = equals(cell.value, "A1")

to_console(printf("%A"))(arg)

